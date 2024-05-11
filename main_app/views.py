import os

from django.shortcuts import render
from django.urls import reverse, reverse_lazy
from django.http import HttpResponseRedirect, HttpResponse
from django_filters.views import FilterView
from openpyxl.workbook import Workbook
from rest_framework import status, request
from rest_framework.response import Response


# import requests


# Create your views here.

#
# def registration(request):
#     a = 'adfghj'
#     return render(request, 'main_app/index.html')


from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
# from django.contrib.auth.models import User
from rest_framework.authtoken.models import Token

from Pract import settings
from api_app.models import *
from .forms import UserRegistrationForm, UserLoginForm
from .myfilters import BrandFilter, CatsFilter, ColorsFilter, SizesFilter, GoodsFilter, HugeCardFilter, MainCatsFilter, \
    MainBrandsFilter, MainProductsFilter, SizeToGoodTableFilter, OrderFilter, OrderToGoodFilter


def redirect_to_data(routing_name, **kwargs):
    import requests
    r = requests.post(f'http://localhost:8000/api/{routing_name}/', data=kwargs)
    return r.text


def register(request):
    from api_app.models import Roles
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('login')
            password = form.cleaned_data.get('password')
            role = Roles.objects.get(name='user')
            user = redirect_to_data('signup', login=username, password=password, role_id=role.id)
            if user is not None:
                return redirect('login_page')  # Перенаправляем на страницу входа

            return Response({'error': 'Please provide both username and password'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        form = UserRegistrationForm()
    return render(request, 'main_app/index.html', {'form': form})


def user_login(request):
    if request.method == 'POST':
        form = UserLoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('login')
            password = form.cleaned_data.get('password')
            user = redirect_to_data('login', login=username, password=password)
            if 'token' in user:
                return redirect('home')  # Перенаправляем на домашнюю страницу
    else:
        form = UserLoginForm()
    return render(request, 'main_app/login.html', {'form': form})


def home_page(request):
    return render(request, 'main_app/home.html')


from django.views.generic import ListView, DetailView, CreateView, DeleteView, UpdateView
# ================================GOODS===================================================================================


class GoodsListView(FilterView, ListView):
    model = Goods
    filterset_class = GoodsFilter
    context_object_name = 'goods_list'
    template_name = 'main_app/goods/ListView.html'
    paginate_by = 4  # Количество объектов на страницу
    queryset = Goods.objects.all().order_by('model_name')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['goods_no_active'] = Goods.objects.all().order_by('model_name')
        context['avalible_brands'] = Brands.objects.all().order_by('brand_name').filter(is_active=True)
        context['avalible_cats'] = Category.objects.all().order_by('category_name').filter(is_active=True)

        return context

    def get(self, request, *args, **kwargs):
        if 'graph' in request.GET:
            return self.create_sales_chart()
        if 'export' in request.GET:
            return self.export_to_excel()
        if 'backup_db' in request.GET:
            os.system(f'python manage.py dumpdata > {settings.BASE_DIR}/backup.json')  # Шаг 1
            with open(f'{settings.BASE_DIR}/backup.json', 'r') as file:  # Шаг 2
                response = HttpResponse(file.read(), content_type='application/json')
                response['Content-Disposition'] = 'attachment; filename=backup.json'  # Шаг 3
                return response

        return super(GoodsListView, self).get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if 'delete_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=False)
        elif 'restore_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=True)
        return HttpResponseRedirect(reverse_lazy('goods_list'))

    def export_to_excel(self):
        goods_queryset = self.get_queryset()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="brands.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Goods'

        columns = [
            'ID', 'Model Name', 'Main pic', 'Category',
            'Brand id', 'Description', 'Price', 'Price with Sale',
            'Color', 'Articul', 'Sale confirmed', 'Is Active'
        ]
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all goods
        for good in goods_queryset:
            row_num += 1
            row = [
                good.pk,
                good.model_name,
                good.main_pic.url,
                good.category,
                good.brand_id,
                good.description,
                good.price,
                good.price_with_sale,
                good.color_id,
                good.articul,
                good.sale_confirmed,
                good.is_active,
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response

    def create_sales_chart(request):
        import matplotlib.pyplot as plt
        from django.http import HttpResponse
        from io import BytesIO
        from django.db.models import Count, Sum
        from django.db.models.functions import TruncDay

        daily_sales = Order.objects.annotate(day=TruncDay('date_added')).values('day').annotate(
            total_sales=Count('id')
        ).order_by('day')

        # Установим настройки для графика
        plt.figure(figsize=(10, 5))
        plt.plot([sale['day'] for sale in daily_sales], [sale['total_sales'] for sale in daily_sales], marker='o')

        # Добавляем заголовки и метки
        plt.title('Статистика продаж по дням')
        plt.xlabel('Дата')
        plt.ylabel('Количество продаж')

        # Сохраняем график в виртуальную файловую систему в памяти
        buffer = BytesIO()
        plt.savefig(buffer, format='png')
        buffer.seek(0)

        # Создаем HTTP-ответ с изображением графика
        response = HttpResponse(buffer.getvalue(), content_type='image/png')

        # Освобождаем ресурсы
        plt.close()
        buffer.close()

        return response


class GoodsDetailView(DetailView):
    model = Goods
    context_object_name = 'goods'
    template_name = 'main_app/goods/DetailView.html'


class GoodsCreateView(CreateView):
    model = Goods
    fields = ['model_name', 'main_pic', 'category', 'brand_id',
              'description', 'price', 'price_with_sale', 'color_id',
              'articul', 'sale_confirmed', 'is_active', ]
    template_name = 'main_app/goods/CreateView.html'
    success_url = reverse_lazy('goods_list')  # Re


class GoodsDeleteView(DeleteView):
    model = Goods
    context_object_name = 'goods'
    template_name = 'main_app/goods/DeleteView.html'
    success_url = reverse_lazy('goods_list')


class GoodsUpdateView(UpdateView):
    model = Goods
    fields = ['model_name', 'main_pic', 'category', 'brand_id',
              'description', 'price', 'price_with_sale', 'color_id',
              'articul', 'sale_confirmed', 'is_active', ]
    template_name = 'main_app/goods/CreateView.html'
    success_url = reverse_lazy('goods_list')


# ================================Brands===================================================================================
class BrandsListView(FilterView, ListView):
    model = Brands
    filterset_class = BrandFilter
    context_object_name = 'brands_list'
    template_name = 'main_app/brands/ListView.html'
    paginate_by = 4  # Количество объектов на страницу
    queryset = Brands.objects.all().order_by('brand_name')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['brands_no_active'] = Brands.objects.all().order_by('brand_name')
        return context

    def get(self, request, *args, **kwargs):
        if 'export' in request.GET:
            return self.export_to_excel()
        if 'backup_db' in request.GET:
            os.system(f'python manage.py dumpdata > {settings.BASE_DIR}/backup.json')  # Шаг 1
            with open(f'{settings.BASE_DIR}/backup.json', 'r') as file:  # Шаг 2
                response = HttpResponse(file.read(), content_type='application/json')
                response['Content-Disposition'] = 'attachment; filename=backup.json'  # Шаг 3
                return response

        return super(BrandsListView, self).get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if 'delete_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=False)
        elif 'restore_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=True)
        return HttpResponseRedirect(reverse_lazy('brands_list'))

    def export_to_excel(self):
        brands_queryset = self.get_queryset()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="brands.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Goods'

        columns = ['ID', 'Brand Name', 'brands Pic', 'Is Active']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all goods
        for brand in brands_queryset:
            row_num += 1
            row = [
                brand.pk,
                brand.brand_name,
                brand.brands_pic.url,
                brand.is_active,
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


class BrandsDetailView(DetailView):
    model = Brands
    context_object_name = 'brands'
    template_name = 'main_app/brands/DetailView.html'


class BrandsCreateView(CreateView):
    model = Brands
    fields = '__all__'
    template_name = 'main_app/brands/CreateView.html'
    success_url = reverse_lazy('brands_list')  # Re


class BrandsDeleteView(DeleteView):
    model = Brands
    context_object_name = 'brands'
    template_name = 'main_app/brands/DeleteView.html'
    success_url = reverse_lazy('brands_list')


class BrandsUpdateView(UpdateView):
    model = Brands
    fields = '__all__'
    template_name = 'main_app/brands/CreateView.html'
    success_url = reverse_lazy('brands_list')


# ================================CTEGORY===================================================================================
class CategoryListView(FilterView, ListView):
    model = Category
    filterset_class = CatsFilter
    context_object_name = 'cats_list'
    template_name = 'main_app/cats/ListView.html'
    paginate_by = 4  # Количество объектов на страницу
    queryset = Category.objects.all().order_by('category_name')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['cats_no_active'] = Category.objects.all().order_by('category_name')
        context['avalible_brands'] = Brands.objects.all().order_by('brand_name').filter(is_active=True)
        return context

    def get(self, request, *args, **kwargs):
        if 'export' in request.GET:
            return self.export_to_excel()
        if 'backup_db' in request.GET:
            os.system(f'python manage.py dumpdata > {settings.BASE_DIR}/backup.json')  # Шаг 1
            with open(f'{settings.BASE_DIR}/backup.json', 'r') as file:  # Шаг 2
                response = HttpResponse(file.read(), content_type='application/json')
                response['Content-Disposition'] = 'attachment; filename=backup.json'  # Шаг 3
                return response

        return super(CategoryListView, self).get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if 'delete_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=False)
        elif 'restore_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=True)
        return HttpResponseRedirect(reverse_lazy('cats_list'))

    def export_to_excel(self):
        cats_queryset = self.get_queryset()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="brands.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Category'

        columns = ['ID', 'Brand Name', 'Category Name', 'Category pic','Is Active']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all goods
        for cat in cats_queryset:
            row_num += 1
            row = [
                cat.pk,
                cat.brand,
                cat.category_name,
                cat.cat_pic.url,
                cat.is_active,
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


class CategoryDetailView(DetailView):
    model = Category
    context_object_name = 'cats'
    template_name = 'main_app/cats/DetailView.html'


class CategoryCreateView(CreateView):
    model = Category
    fields = '__all__'
    template_name = 'main_app/cats/CreateView.html'
    success_url = reverse_lazy('cats_list')  # Re


class CategoryDeleteView(DeleteView):
    model = Category
    context_object_name = 'cats'
    template_name = 'main_app/cats/DeleteView.html'
    success_url = reverse_lazy('cats_list')


class CategoryUpdateView(UpdateView):
    model = Category
    fields = '__all__'
    template_name = 'main_app/cats/CreateView.html'
    success_url = reverse_lazy('cats_list')

# ================================COLORS===================================================================================
class ColorsListView(FilterView, ListView):
    model = Colors
    filterset_class = ColorsFilter
    context_object_name = 'colors_list'
    template_name = 'main_app/colors/ListView.html'
    paginate_by = 4  # Количество объектов на страницу
    queryset = Colors.objects.all().order_by('color')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['colors_no_active'] = Colors.objects.all().order_by('color')
        return context

    def get(self, request, *args, **kwargs):
        if 'export' in request.GET:
            return self.export_to_excel()
        if 'backup_db' in request.GET:
            os.system(f'python manage.py dumpdata > {settings.BASE_DIR}/backup.json')  # Шаг 1
            with open(f'{settings.BASE_DIR}/backup.json', 'r') as file:  # Шаг 2
                response = HttpResponse(file.read(), content_type='application/json')
                response['Content-Disposition'] = 'attachment; filename=backup.json'  # Шаг 3
                return response

        return super(ColorsListView, self).get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if 'delete_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=False)
        elif 'restore_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=True)
        return HttpResponseRedirect(reverse_lazy('colors_list'))

    def export_to_excel(self):
        colors_queryset = self.get_queryset()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="brands.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Goods'

        columns = ['ID', 'Color', 'Is Active']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all goods
        for color in colors_queryset:
            row_num += 1
            row = [
                color.pk,
                color.color,
                color.is_active,
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


class ColorsDetailView(DetailView):
    model = Colors
    context_object_name = 'colors'
    template_name = 'main_app/colors/DetailView.html'


class ColorsCreateView(CreateView):
    model = Colors
    fields = '__all__'
    template_name = 'main_app/colors/CreateView.html'
    success_url = reverse_lazy('colors_list')  # Re


class ColorsDeleteView(DeleteView):
    model = Colors
    context_object_name = 'colors'
    template_name = 'main_app/colors/DeleteView.html'
    success_url = reverse_lazy('colors_list')


class ColorsUpdateView(UpdateView):
    model = Colors
    fields = '__all__'
    template_name = 'main_app/colors/CreateView.html'
    success_url = reverse_lazy('colors_list')
# ================================SIZES===================================================================================
class SizesListView(FilterView, ListView):
    model = Sizes
    filterset_class = SizesFilter
    context_object_name = 'sizes_list'
    template_name = 'main_app/sizes/ListView.html'
    paginate_by = 4  # Количество объектов на страницу
    queryset = Sizes.objects.all().order_by('size')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sizes_no_active'] = Sizes.objects.all().order_by('size')
        return context

    def get(self, request, *args, **kwargs):
        if 'export' in request.GET:
            return self.export_to_excel()
        if 'backup_db' in request.GET:
            os.system(f'python manage.py dumpdata > {settings.BASE_DIR}/backup.json')  # Шаг 1
            with open(f'{settings.BASE_DIR}/backup.json', 'r') as file:  # Шаг 2
                response = HttpResponse(file.read(), content_type='application/json')
                response['Content-Disposition'] = 'attachment; filename=backup.json'  # Шаг 3
                return response

        return super(SizesListView, self).get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if 'delete_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=False)
        elif 'restore_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=True)
        return HttpResponseRedirect(reverse_lazy('sizes_list'))

    def export_to_excel(self):
        colors_queryset = self.get_queryset()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="brands.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Sizes'

        columns = ['ID', 'Size', 'Is Active']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all goods
        for color in colors_queryset:
            row_num += 1
            row = [
                color.pk,
                color.size,
                color.is_active,
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


class SizesDetailView(DetailView):
    model = Sizes
    context_object_name = 'sizes'
    template_name = 'main_app/sizes/DetailView.html'


class SizesCreateView(CreateView):
    model = Sizes
    fields = '__all__'
    template_name = 'main_app/sizes/CreateView.html'
    success_url = reverse_lazy('sizes_list')  # Re


class SizesDeleteView(DeleteView):
    model = Sizes
    context_object_name = 'sizes'
    template_name = 'main_app/sizes/DeleteView.html'
    success_url = reverse_lazy('sizes_list')


class SizesUpdateView(UpdateView):
    model = Sizes
    fields = '__all__'
    template_name = 'main_app/sizes/CreateView.html'
    success_url = reverse_lazy('sizes_list')
# ================================HUGECARD===================================================================================
class HugeCardListView(FilterView, ListView):
    model = HugeCard
    filterset_class = HugeCardFilter
    context_object_name = 'hugecards_list'
    template_name = 'main_app/hugecards/ListView.html'
    paginate_by = 4  # Количество объектов на страницу
    queryset = HugeCard.objects.all().order_by('description')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['hugecards_no_active'] = HugeCard.objects.all().order_by('description')
        context['avalible_goods'] = Goods.objects.all().order_by('model_name').filter(is_active=True)
        return context

    def get(self, request, *args, **kwargs):
        if 'export' in request.GET:
            return self.export_to_excel()
        if 'backup_db' in request.GET:
            os.system(f'python manage.py dumpdata > {settings.BASE_DIR}/backup.json')  # Шаг 1
            with open(f'{settings.BASE_DIR}/backup.json', 'r') as file:  # Шаг 2
                response = HttpResponse(file.read(), content_type='application/json')
                response['Content-Disposition'] = 'attachment; filename=backup.json'  # Шаг 3
                return response

        return super(HugeCardListView, self).get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if 'delete_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=False)
        elif 'restore_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=True)
        return HttpResponseRedirect(reverse_lazy('hugecards_list'))

    def export_to_excel(self):
        hugecards_queryset = self.get_queryset()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="brands.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Category'

        columns = ['ID', 'Good Id', 'Description', 'Is Active']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all goods
        for card in hugecards_queryset:
            row_num += 1
            row = [
                card.pk,
                card.good_id,
                card.description,
                card.is_active,
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


class HugeCardDetailView(DetailView):
    model = HugeCard
    context_object_name = 'hugecards'
    template_name = 'main_app/hugecards/DetailView.html'


class HugeCardCreateView(CreateView):
    model = HugeCard
    fields = '__all__'
    template_name = 'main_app/hugecards/CreateView.html'
    success_url = reverse_lazy('hugecards_list')  # Re


class HugeCardDeleteView(DeleteView):
    model = HugeCard
    context_object_name = 'hugecards'
    template_name = 'main_app/hugecards/DeleteView.html'
    success_url = reverse_lazy('hugecards_list')


class HugeCardUpdateView(UpdateView):
    model = HugeCard
    fields = '__all__'
    template_name = 'main_app/hugecards/CreateView.html'
    success_url = reverse_lazy('hugecards_list')
# ================================MAINCATS===================================================================================
class MainCatsListView(FilterView, ListView):
    model = MainCats
    filterset_class = MainCatsFilter
    context_object_name = 'maincats_list'
    template_name = 'main_app/maincats/ListView.html'
    paginate_by = 4  # Количество объектов на страницу
    queryset = MainCats.objects.all().order_by('cat_id__category_name')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['maincats_no_active'] = MainCats.objects.all().order_by('cat_id__category_name')
        context['avalible_cats'] = Category.objects.all().order_by('category_name').filter(is_active=True)
        return context

    def get(self, request, *args, **kwargs):
        if 'export' in request.GET:
            return self.export_to_excel()
        if 'backup_db' in request.GET:
            os.system(f'python manage.py dumpdata > {settings.BASE_DIR}/backup.json')  # Шаг 1
            with open(f'{settings.BASE_DIR}/backup.json', 'r') as file:  # Шаг 2
                response = HttpResponse(file.read(), content_type='application/json')
                response['Content-Disposition'] = 'attachment; filename=backup.json'  # Шаг 3
                return response

        return super(MainCatsListView, self).get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if 'delete_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=False)
        elif 'restore_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=True)
        return HttpResponseRedirect(reverse_lazy('maincats_list'))

    def export_to_excel(self):
        main_cats_queryset = self.get_queryset()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="brands.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'MainCategory'

        columns = ['ID', 'Category Id', 'Is Active']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all goods
        for card in main_cats_queryset:
            row_num += 1
            row = [
                card.pk,
                card.cat_id,
                card.is_active,
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


class MainCatsDetailView(DetailView):
    model = MainCats
    context_object_name = 'maincats'
    template_name = 'main_app/maincats/DetailView.html'


class MainCatsCreateView(CreateView):
    model = MainCats
    fields = '__all__'
    template_name = 'main_app/maincats/CreateView.html'
    success_url = reverse_lazy('maincats_list')  # Re


class MainCatsDeleteView(DeleteView):
    model = MainCats
    context_object_name = 'maincats'
    template_name = 'main_app/maincats/DeleteView.html'
    success_url = reverse_lazy('maincats_list')


class MainCatsUpdateView(UpdateView):
    model = MainCats
    fields = '__all__'
    template_name = 'main_app/maincats/CreateView.html'
    success_url = reverse_lazy('maincats_list')
# ================================MAINBRANDS===================================================================================
class MainBrandsListView(FilterView, ListView):
    model = MainBrands
    filterset_class = MainBrandsFilter
    context_object_name = 'mainbrands_list'
    template_name = 'main_app/mainbrands/ListView.html'
    paginate_by = 4  # Количество объектов на страницу
    queryset = MainBrands.objects.all().order_by('brand_id__brand_name')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['mainbrands_no_active'] = MainBrands.objects.all().order_by('brand_id__brand_name')
        context['avalible_brands'] = Brands.objects.all().order_by('brand_name').filter(is_active=True)
        return context

    def get(self, request, *args, **kwargs):
        if 'export' in request.GET:
            return self.export_to_excel()
        if 'backup_db' in request.GET:
            os.system(f'python manage.py dumpdata > {settings.BASE_DIR}/backup.json')  # Шаг 1
            with open(f'{settings.BASE_DIR}/backup.json', 'r') as file:  # Шаг 2
                response = HttpResponse(file.read(), content_type='application/json')
                response['Content-Disposition'] = 'attachment; filename=backup.json'  # Шаг 3
                return response

        return super(MainBrandsListView, self).get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if 'delete_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=False)
        elif 'restore_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=True)
        return HttpResponseRedirect(reverse_lazy('mainbrands_list'))

    def export_to_excel(self):
        main_brands_queryset = self.get_queryset()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="brands.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'MainCategory'

        columns = ['ID', 'Brand Id', 'Is Active']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all goods
        for brand in main_brands_queryset:
            row_num += 1
            row = [
                brand.pk,
                brand.brand_id,
                brand.is_active,
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


class MainBrandsDetailView(DetailView):
    model = MainBrands
    context_object_name = 'mainbrands'
    template_name = 'main_app/mainbrands/DetailView.html'


class MainBrandsCreateView(CreateView):
    model = MainBrands
    fields = '__all__'
    template_name = 'main_app/mainbrands/CreateView.html'
    success_url = reverse_lazy('mainbrands_list')  # Re


class MainBrandsDeleteView(DeleteView):
    model = MainBrands
    context_object_name = 'mainbrands'
    template_name = 'main_app/mainbrands/DeleteView.html'
    success_url = reverse_lazy('mainbrands_list')


class MainBrandsUpdateView(UpdateView):
    model = MainBrands
    fields = '__all__'
    template_name = 'main_app/mainbrands/CreateView.html'
    success_url = reverse_lazy('mainbrands_list')


# ================================MAINPRODUCTS===================================================================================
class MainProductsListView(FilterView, ListView):
    model = MainProducts
    filterset_class = MainProductsFilter
    context_object_name = 'mainproducts_list'
    template_name = 'main_app/mainproducts/ListView.html'
    paginate_by = 4  # Количество объектов на страницу
    queryset = MainProducts.objects.all().order_by('good_id__model_name')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['mainproducts_no_active'] = MainProducts.objects.all().order_by('good_id__model_name')
        context['avalible_goods'] = Goods.objects.all().order_by('model_name').filter(is_active=True)
        return context

    def get(self, request, *args, **kwargs):
        if 'export' in request.GET:
            return self.export_to_excel()
        if 'backup_db' in request.GET:
            os.system(f'python manage.py dumpdata > {settings.BASE_DIR}/backup.json')  # Шаг 1
            with open(f'{settings.BASE_DIR}/backup.json', 'r') as file:  # Шаг 2
                response = HttpResponse(file.read(), content_type='application/json')
                response['Content-Disposition'] = 'attachment; filename=backup.json'  # Шаг 3
                return response

        return super(MainProductsListView, self).get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if 'delete_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=False)
        elif 'restore_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=True)
        return HttpResponseRedirect(reverse_lazy('mainbrands_list'))

    def export_to_excel(self):
        main_products_queryset = self.get_queryset()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="brands.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'MainCategory'

        columns = ['ID', 'Good Id', 'Is Active']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all goods
        for good in main_products_queryset:
            row_num += 1
            row = [
                good.pk,
                good.good_id,
                good.is_active,
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


class MainProductsDetailView(DetailView):
    model = MainProducts
    context_object_name = 'mainproducts'
    template_name = 'main_app/mainproducts/DetailView.html'


class MainProductsCreateView(CreateView):
    model = MainProducts
    fields = '__all__'
    template_name = 'main_app/mainproducts/CreateView.html'
    success_url = reverse_lazy('mainproducts_list')  # Re


class MainProductsDeleteView(DeleteView):
    model = MainProducts
    context_object_name = 'mainproducts'
    template_name = 'main_app/mainproducts/DeleteView.html'
    success_url = reverse_lazy('mainproducts_list')


class MainProductsUpdateView(UpdateView):
    model = MainProducts
    fields = '__all__'
    template_name = 'main_app/mainproducts/CreateView.html'
    success_url = reverse_lazy('mainproducts_list')


# ================================SIZETOGOOD===================================================================================
class SizeToGoodTableListView(FilterView, ListView):
    model = SizesToGoodTable
    filterset_class = SizeToGoodTableFilter
    context_object_name = 'sizestogoodtable_list'
    template_name = 'main_app/sizestogoodtable/ListView.html'
    paginate_by = 4  # Количество объектов на страницу
    queryset = SizesToGoodTable.objects.all().order_by('good__model_name')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sizestogood_no_active'] = SizesToGoodTable.objects.all().order_by('good__model_name')
        context['avalible_goods'] = Goods.objects.all().order_by('model_name').filter(is_active=True)
        context['avalible_sizes'] = Sizes.objects.all().order_by('size')
        return context

    def get(self, request, *args, **kwargs):
        if 'export' in request.GET:
            return self.export_to_excel()
        if 'backup_db' in request.GET:
            os.system(f'python manage.py dumpdata > {settings.BASE_DIR}/backup.json')  # Шаг 1
            with open(f'{settings.BASE_DIR}/backup.json', 'r') as file:  # Шаг 2
                response = HttpResponse(file.read(), content_type='application/json')
                response['Content-Disposition'] = 'attachment; filename=backup.json'  # Шаг 3
                return response

        return super(SizeToGoodTableListView, self).get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if 'delete_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=False)
        elif 'restore_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=True)
        return HttpResponseRedirect(reverse_lazy('mainbrands_list'))

    def export_to_excel(self):
        sizes_to_good_queryset = self.get_queryset()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="brands.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'MainCategory'

        columns = ['ID', 'Good', 'Size', 'Count','Is Active']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all goods
        for zap in sizes_to_good_queryset:
            row_num += 1
            row = [
                zap.pk,
                zap.good.model_name,
                zap.size.size,
                zap.count,
                zap.is_active,
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


class SizeToGoodTableDetailView(DetailView):
    model = SizesToGoodTable
    context_object_name = 'sizestogoodtable'
    template_name = 'main_app/sizestogoodtable/DetailView.html'


class SizeToGoodTableCreateView(CreateView):
    model = SizesToGoodTable
    fields = '__all__'
    template_name = 'main_app/sizestogoodtable/CreateView.html'
    success_url = reverse_lazy('sizestogoodtable_list')  # Re


class SizeToGoodTableDeleteView(DeleteView):
    model = SizesToGoodTable
    context_object_name = 'sizestogoodtable'
    template_name = 'main_app/sizestogoodtable/DeleteView.html'
    success_url = reverse_lazy('sizestogoodtable_list')


class SizeToGoodTableUpdateView(UpdateView):
    model = SizesToGoodTable
    fields = '__all__'
    template_name = 'main_app/sizestogoodtable/CreateView.html'
    success_url = reverse_lazy('sizestogoodtable_list')


# ================================ORDER===================================================================================
class OrderListView(FilterView, ListView):
    model = Order
    filterset_class = OrderFilter
    context_object_name = 'order_list'
    template_name = 'main_app/order/ListView.html'
    paginate_by = 4  # Количество объектов на страницу
    queryset = Order.objects.all().order_by('user__login')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['orders_no_active'] = Order.objects.all().order_by('user__login')
        context['avalible_users'] = Users.objects.all().order_by('login')
        return context

    def get(self, request, *args, **kwargs):
        if 'export' in request.GET:
            return self.export_to_excel()
        if 'backup_db' in request.GET:
            os.system(f'python manage.py dumpdata > {settings.BASE_DIR}/backup.json')  # Шаг 1
            with open(f'{settings.BASE_DIR}/backup.json', 'r') as file:  # Шаг 2
                response = HttpResponse(file.read(), content_type='application/json')
                response['Content-Disposition'] = 'attachment; filename=backup.json'  # Шаг 3
                return response

        return super(OrderListView, self).get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if 'delete_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=False)
        elif 'restore_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=True)
        return HttpResponseRedirect(reverse_lazy('order_list'))

    def export_to_excel(self):
        sizes_to_good_queryset = self.get_queryset()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="brands.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'MainCategory'

        columns = ['ID', 'User', 'Data', 'Phone', 'Is Active', 'Is Paid']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all goods
        for zap in sizes_to_good_queryset:
            row_num += 1
            row = [
                zap.pk,
                zap.user,
                zap.date_added,
                zap.is_active,
                zap.is_paid
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


class OrderDetailView(DetailView):
    model = Order
    context_object_name = 'order'
    template_name = 'main_app/order/DetailView.html'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['goods_for_user'] = self.get_object().goodtouser_set.all()
        return context


class OrderCreateView(CreateView):
    model = Order
    fields = ['user', 'phone_number', 'is_active', 'is_paid']
    template_name = 'main_app/order/CreateView.html'
    success_url = reverse_lazy('order_list')  # Re


class OrderDeleteView(DeleteView):
    model = Order
    context_object_name = 'order'
    template_name = 'main_app/order/DeleteView.html'
    success_url = reverse_lazy('order_list')


class OrderUpdateView(UpdateView):
    model = Order
    fields = ['user', 'phone_number', 'is_active', 'is_paid']
    template_name = 'main_app/order/CreateView.html'
    success_url = reverse_lazy('order_list')
# ================================GOODTOUSER===================================================================================

class GoodToListView(FilterView, ListView):
    model = GoodToUser
    filterset_class = OrderToGoodFilter
    context_object_name = 'ordertouser_list'
    template_name = 'main_app/ordertouser/ListView.html'
    paginate_by = 4  # Количество объектов на страницу
    queryset = GoodToUser.objects.all().order_by('order__user__login')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['ordertogood_no_active'] = GoodToUser.objects.all().order_by('order__user__login')
        context['avalible_goods'] = Goods.objects.all().order_by('model_name')
        context['avalible_orders'] = Order.objects.all().order_by('user__login')
        return context

    def get(self, request, *args, **kwargs):
        if 'export' in request.GET:
            return self.export_to_excel()
        if 'backup_db' in request.GET:
            os.system(f'python manage.py dumpdata > {settings.BASE_DIR}/backup.json')  # Шаг 1
            with open(f'{settings.BASE_DIR}/backup.json', 'r') as file:  # Шаг 2
                response = HttpResponse(file.read(), content_type='application/json')
                response['Content-Disposition'] = 'attachment; filename=backup.json'  # Шаг 3
                return response

        return super(GoodToListView, self).get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if 'delete_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=False)
        elif 'restore_selected' in request.POST:
            self.model.objects.filter(id__in=request.POST.getlist('selected_ids')).update(is_active=True)
        return HttpResponseRedirect(reverse_lazy('ordertouser_list'))

    def export_to_excel(self):
        sizes_to_good_queryset = self.get_queryset()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="brands.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'MainCategory'

        columns = ['ID', 'User', 'Good', 'Is Active']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all goods
        for zap in sizes_to_good_queryset:
            row_num += 1
            row = [
                zap.pk,
                zap.user,
                zap.good,
                zap.is_active,
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


class GoodToDetailView(DetailView):
    model = GoodToUser
    context_object_name = 'ordertouser'
    template_name = 'main_app/ordertouser/DetailView.html'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['goods_for_user'] = self.get_object().goodtouser_set.all()
        return context


class GoodToCreateView(CreateView):
    model = GoodToUser
    fields = '__all__'
    template_name = 'main_app/ordertouser/CreateView.html'
    success_url = reverse_lazy('ordertouser_list')  # Re


class GoodToDeleteView(DeleteView):
    model = GoodToUser
    context_object_name = 'order'
    template_name = 'main_app/ordertouser/DeleteView.html'
    success_url = reverse_lazy('ordertouser_list')


class GoodToUpdateView(UpdateView):
    model = GoodToUser
    fields = '__all__'
    template_name = 'main_app/ordertouser/CreateView.html'
    success_url = reverse_lazy('ordertouser_list')



